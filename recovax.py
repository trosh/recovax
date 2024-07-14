#! /usr/bin/env python3

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QVBoxLayout,
    QHBoxLayout,
    QScrollArea,
    QCheckBox,
    QLineEdit,
    QLabel,
    QPushButton,
    QWidget,
)

from openpyxl import load_workbook

wb = load_workbook(filename = "vaccins.xlsx")
ws = wb["Data"]
# Sanity check
assert(ws["A1"].value.startswith("Vaccin"))
assert(ws["B1"].value.startswith("Âge"))
assert(ws["C1"].value.startswith("Conditions"))
assert(ws["D1"].value.startswith("Description"))

app = QApplication([])

class MainWindow(QMainWindow):

    def __init__(self):
        super(MainWindow, self).__init__()
        self.setWindowTitle("[v0.2] Aide à la recommandation vaccinale")
        self.construire_conditions()
        layout = QVBoxLayout()
        layout.addWidget(QLabel("<b style='font-size: 3em; color: red'>Ceci est un outil en développement, à ne pas utiliser en contexte médical</b>"))
        # Add widgets
        age_label = QLabel("Âge du patient :")
        age = QLineEdit()
        age.setMaxLength(7)
        age.setPlaceholderText("x an(s) / y mois")
        age_layout = QHBoxLayout()
        age_layout.addWidget(age_label)
        age_layout.addWidget(age)
        age_line = QWidget()
        age_line.setLayout(age_layout)
        layout.addWidget(age_line)
        conds_layout = QVBoxLayout()
        self.conds = list()
        for c in self.conditions:
            if c == "toute comorbidité":
                continue
            checkbox = QCheckBox()
            checkbox.setCheckState(Qt.Unchecked)
            checkbox.setFixedWidth(16)
            cond_label = QLabel(c)
            cond_label.setMaximumWidth(1000)
            cond_label.setWordWrap(True)
            self.conds.append([checkbox, cond_label])
            cond_layout = QHBoxLayout()
            cond_layout.addWidget(checkbox)
            cond_layout.addWidget(cond_label)
            cond_line = QWidget()
            cond_line.setLayout(cond_layout)
            conds_layout.addWidget(cond_line)
        conds_widget = QWidget()
        conds_widget.setLayout(conds_layout)
        conds_scroll = QScrollArea()
        conds_scroll.setWidget(conds_widget)
        layout.addWidget(conds_scroll)
        # Bouton envoi
        bouton = QPushButton("Recommandations")
        bouton.clicked.connect(self.envoi)
        layout.addWidget(bouton)
        # Finish setting up page
        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)
        # References
        self.layout = layout
        self.age = age

    def envoi(self):
        patient = dict()
        age_split = self.age.text().split(" ")
        if len(age_split) != 2:
            print("TODO dialog")
            return
        if age_split[1].lower() not in ["an", "ans"]:
            # TODO mois (bébés)
            print("TODO dialog")
            return
        patient["age"] = int(age_split[0])
        patient["conditions"] = list()
        for (checkbox, label) in self.conds:
            if checkbox.checkState() == Qt.Checked:
                patient["conditions"].append(label.text())
                print(label.text())
        self.recommandations(patient)

    def construire_conditions(self):
        self.conditions = list()
        for regle in ws.iter_rows(min_row=2):
            if regle[2] is None or regle[2].value is None:
                continue
            for c in regle[2].value.split(";"):
                c = c.strip()
                if c == "rattrapage":
                    continue
                if c not in self.conditions:
                    self.conditions.append(c)

    def recommandations(self, patient):
        regles_applicables = list()
        for regle in ws.iter_rows(min_row=2):
            if regle[0] is None or regle[0].value is None:
                continue
            # Vaccin contre
            vaccin_contre = regle[0].value
            # Âge
            if regle[1] is None or regle[1].value is None:
                age = None
            else:
                age_str = regle[1].value
                if age_str.endswith("+"):
                    age = int(age_str[:-1])
                else:
                    ages = age_str.split("-", 2)
                    age = (int(ages[0]), int(ages[1]))
            # Condition
            if regle[2] is None or regle[2].value is None:
                conditions = None
            else:
                conditions = regle[2].value
                if conditions is not None:
                    conditions = list(map(str.strip, conditions.split(";")))
            # Description, précisions
            if regle[3] is None or regle[3].value is None:
                description = None
            else:
                description = regle[3].value
            ###### TRAITEMENT DE LA REGLE ######
            if age is not None:
                if type(age) is int:
                    if patient["age"] < age:
                        continue
                elif patient["age"] < age[0] \
                  or patient["age"] > age[1]:
                    continue
            regles_applicables.append({
                "vaccin_contre": vaccin_contre,
                "age": age,
                "conditions": conditions,
                "description": description,
            })
        self.clearLayout(self.layout)
        self.layout.addWidget(QLabel("<b>Vaccinations :</b><br>"))
        deja_faits = list()
        for regle in regles_applicables:
            conditions = regle["conditions"]
            if conditions is not None:
                match_all = True
                for c in conditions:
                    if c == "rattrapage" \
                    or not ((c == "toute comorbidité" and len(patient["conditions"]) > 0) \
                            or c in patient["conditions"]):
                        match_all = False
                        break
                if not match_all:
                    continue
            if regle["vaccin_contre"] in deja_faits:
                continue
            deja_faits.append(regle["vaccin_contre"])
            self.affichage_vaccin(
                regle["vaccin_contre"],
                regle["age"],
                regle["description"])
        self.layout.addWidget(QLabel("<b>Rattrapages :</b><br>"))
        for regle in regles_applicables:
            conditions = regle["conditions"]
            if conditions is None:
                continue
            if "rattrapage" in conditions:
                match_all = True
                for c in conditions:
                    if c == "rattrapage":
                        continue
                    if c not in patient["conditions"]:
                        match_all = False
                        break
                if not match_all:
                    continue
                self.affichage_vaccin(
                    regle["vaccin_contre"],
                    regle["age"],
                    regle["description"])

    def affichage_vaccin(self, vaccin_contre, age, description):
        self.layout.addWidget(QLabel(f"- Vaccin contre {vaccin_contre}"))
        if age is not None:
            if type(age) is int:
                self.layout.addWidget(QLabel(f"\t{age} ans et plus"))
            else:
                self.layout.addWidget(QLabel(f"\tentre {age[0]} et {age[1]} ans"))
        if description is not None:
            self.layout.addWidget(QLabel(f"\t{description}"))
        self.layout.addWidget(QLabel(""))

    def clearLayout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clearLayout(item.layout())

window = MainWindow()
window.show()

app.exec()
